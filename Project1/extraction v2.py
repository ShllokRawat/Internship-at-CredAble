#ALL THE MODULES ARE IMPORTED HERE
import openpyxl
import xlsxwriter
import re
from datetime import datetime
#ALL THE MODULES ARE IMPORTED HERE

#ALL THE AUTOMATION VARIABLES ARE STORED HERE
x = int(1)
y = int(x +1)
n = int(2)
v = int(2)
u = int(2)
x1 = int(x + 2)
x2 = int(x + 3)
x3 = int(x + 4)
x4 = int(x + 5)
x5 = int(x + 6)
x6 = int(x + 7)
x7 = int(x + 8)
x8 = int(x + 9)
x9 = int(x + 10)
#ALL THE AUTOMATION VARIABLES ARE STORED HERE

#DATE TIME CODE HERE
now = datetime.now()
dt_string = now.strftime("%d-%m-%Y %H.%M.%S")
#DATE TIME CODE HERE
workbook = xlsxwriter.Workbook('BSA analysis report ' + str(dt_string) + '.xlsx')
worksheet = workbook.add_worksheet()
num = int(input("Enter number of files"))
for i in range(0, num):
    path = str(input("Enter the path"))
    path1 = re.sub('"' , '', path)
    firstPass = re.findall("\d+_",path1)[0]
    AppId = re.sub("_","",firstPass)
    wb = openpyxl.load_workbook(path1)
    wb.active = wb['Analysis']
    sheet = wb.active 
    Name = sheet.cell(row = 2, column = 2)
    k = int(2)
    AmountBusCred = int(0)
    for j in range(0, 12):
        tally = sheet.cell(row = 20, column = k)
        AmountBusCred = AmountBusCred + tally.value
        k = k + 1
    Aa = str("A" + str(x))
    Bb = str("B" + str(x))
    Cc = str("C" + str(x))
    A1 = str("A" + str(x1))
    A2 = str("A" + str(x2))
    A3 = str("A" + str(x3))
    A4 = str("A" + str(x4))
    A5 = str("A" + str(x5))
    A6 = str("A" + str(x6))
    A7 = str("A" + str(x7))
    A8 = str("A" + str(x8))
    A9 = str("A" + str(x9))
    B1 = str("B" + str(x1))
    B2 = str("B" + str(x2))
    B3 = str("B" + str(x3))
    B4 = str("B" + str(x4))
    B5 = str("B" + str(x5))
    B6 = str("B" + str(x6))
    B7 = str("B" + str(x7))
    B8 = str("B" + str(x8))
    B9 = str("B" + str(x9))
    C1 = str("C" + str(x1))
    C2 = str("C" + str(x2))
    C3 = str("C" + str(x3))
    C4 = str("C" + str(x4))
    C5 = str("C" + str(x5))
    C6 = str("C" + str(x6))
    C7 = str("C" + str(x7))
    C8 = str("C" + str(x8))
    C9 = str("C" + str(x9))
    Dd = str("D" + str(x))
    Aaa = str("A" + str(y))
    Bbb = str("B" + str(y))
    Ccc = str("C" + str(y))
    Ss = str("S" + str(x))
    Tt = str("T" + str(x))
    worksheet.write(Aa, 'Application ID')
    worksheet.write(Bb, 'Name Of Account Holder')
    worksheet.write(Cc, 'Total Amount of Business Credits')
    worksheet.write(Dd, 'Narration')
    worksheet.write(Aaa, str(AppId))
    worksheet.write(A1, str(AppId))
    worksheet.write(A2, str(AppId))
    worksheet.write(A3, str(AppId))
    worksheet.write(A4, str(AppId))
    worksheet.write(A5, str(AppId))
    worksheet.write(A6, str(AppId))
    worksheet.write(A7, str(AppId))
    worksheet.write(A8, str(AppId))
    worksheet.write(A9, str(AppId))
    worksheet.write(Ss, 'Buyer')
    worksheet.write(Tt, 'Dependency')
    worksheet.write(Bbb, Name.value)
    worksheet.write(B1, str(Name.value))
    worksheet.write(B2, str(Name.value))
    worksheet.write(B3, str(Name.value))
    worksheet.write(B4, str(Name.value))
    worksheet.write(B5, str(Name.value))
    worksheet.write(B6, str(Name.value))
    worksheet.write(B7, str(Name.value))
    worksheet.write(B8, str(Name.value))
    worksheet.write(B9, str(Name.value))
    worksheet.write(Ccc, AmountBusCred)
    worksheet.write(C1, str(AmountBusCred))
    worksheet.write(C2, str(AmountBusCred))
    worksheet.write(C3, str(AmountBusCred))
    worksheet.write(C4, str(AmountBusCred))
    worksheet.write(C5, str(AmountBusCred))
    worksheet.write(C6, str(AmountBusCred))
    worksheet.write(C7, str(AmountBusCred))
    worksheet.write(C8, str(AmountBusCred))
    worksheet.write(C9, str(AmountBusCred))
    wb.active = wb['Top 10 Party Xns1']
    sheet1 = wb.active
    l = int(2)
    for l in range(2, 13):
        for m in range(2, 17):
            cell_obj = sheet1.cell(row = l, column = m)
            if m==2:
                Zz = str("D" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==3:
                Zz = str("E" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==4:
                Zz = str("F" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==5:
                Zz = str("G" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==6:
                Zz = str("H" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==7:
                Zz = str("I" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==8:
                Zz = str("J" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==9:
                Zz = str("K" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==10:
                Zz = str("L" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==11:
                Zz = str("M" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==12:
                Zz = str("N" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==13:
                Zz = str("O" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==14:
                Zz = str("P" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
            elif m==15:
                Zz = str("Q" + str(n-1))
                if l==2:
                    worksheet.write(Zz, cell_obj.value)
                else:
                    Totalavg = int(0)
                    for j in range(3,15):
                        Num = sheet1.cell(row = l, column = j)
                        number = int(Num.value)
                        Totalavg = int(Totalavg + number)
                    worksheet.write(Zz, Totalavg)
            elif m==16:
                Zz = str("R" + str(n-1))
                worksheet.write(Zz, cell_obj.value)
                n = n+1
    
    for g in range(3,13):
        Bbz = str('S' + str(v))
        Buyer = sheet1.cell(row = g, column = 2)
        nx = re.sub('Transfer from ', '', str(Buyer.value))
        worksheet.write(Bbz, nx)
        v = int(v+1)
        
    #THIS IS A CODE FOR THE DEPENDANCY
    for a in range(3,13):
        Tt = str('T'+str(u))
        Totalavg = float(0)
        for j in range(3,15):
            Num = sheet1.cell(row = a, column = j)
            Totalavg = float(Totalavg + Num.value)
        Dependancy = Totalavg/AmountBusCred
        Dependancypercent = round(Dependancy*100, 2)
        worksheet.write(Tt, str(str(Dependancypercent)+'%'))
        u = int(u + 1)

    #AUTOMATION VARIABLES ARE CHANGED TO AUTOMATE HERE
    x = int(x + 12)
    x1 = int(x1 + 12)
    x2 = int(x2 + 12)
    x3 = int(x3 + 12)
    x4 = int(x4 + 12)
    x5 = int(x5 + 12)
    x6 = int(x6 + 12)
    x7 = int(x7 + 12)
    x8 = int(x8 + 12)
    x9 = int(x9 + 12)
    n = int(n + 1)
    y = int(y + 12)
    v = int(v + 2)
    u = int(u + 2)
    #AUTOMATION VARIABLES ARE CHANGED TO AUTOMATE HERE
workbook.close()
    
    
 
    
    
