#ALL THE MODULES ARE IMPORTED HERE
from difflib import SequenceMatcher
import openpyxl
import xlsxwriter
import re
from datetime import datetime
#ALL THE MODULES ARE IMPORTED HERE

workbook = xlsxwriter.Workbook("Percentage Matchv4.xlsx")
wb = openpyxl.load_workbook("C:\\Users\\shllo\\Desktop\\Internship\\Data\\Final Data\\BSA analysis report 28-06-2022 11.27.42.xlsx")
worksheet = workbook.add_worksheet()
wb.active = wb['Sheet1']
wb2 = openpyxl.load_workbook("C:\\Users\\shllo\\Desktop\\Internship\\Data\\Final Data\\Buyer list .xlsx")
wb2.active = wb2['Sheet1']
sheet2 = wb2.active
sheet = wb.active
x = int(2)
z = int(0)
b = int(1)
while(x!=3762):
	if(b == 1):
		
		worksheet.write('U1', 'Percentage Match')
		worksheet.write('V1', 'Buyer Matched')
		b = int(b+1)
	
	Uu = str('U' + str(b))
	Vv = str('V' + str(b))
	a = int(1)
	percentagematch = int(0)
	nameofbuyer = sheet.cell(row = x, column = 19)
	if(nameofbuyer.value == None):
		b = int(b+1)
		
		Uu = str('U' + str(b))
		Vv = str('V' + str(b))
		x = int(x+1)
		buyername2 = sheet.cell(row = x, column = 19)
		if(buyername2.value == 'Buyer'):
			x = int(x+1)
			
			worksheet.write(Uu, 'Percentage Match')
			worksheet.write(Vv, 'Buyer Matched')
			b = int(b+1)
			continue
	buyername = str(nameofbuyer.value)
	buyernamecatching = re.sub('PRIVATE' , '', buyername)
	buyernamecatching2 = re.sub('LIMITED', '', buyernamecatching)
	buyernamecatching3 = re.sub('Private', '', buyernamecatching2)
	buyernamecatching4 = re.sub('Limited', '', buyernamecatching3)
	buyernamecatching5 = re.sub('CONSTRUCTION', '', buyernamecatching4)
	buyernamecatching6 = re.sub('ENTERPRISES', '', buyernamecatching5)
	buyernamecatching7 = re.sub('TRADERS', '', buyernamecatching6)
	buyernamecatching8 = re.sub('INDUSTRIES', '', buyernamecatching7)
	buyernamecatching9 = re.sub('SERVICES', '', buyernamecatching8)
	buyernamecatching10 = re.sub('TECHNOLOGIES', '', buyernamecatching9)
	buyernamecatching11 = re.sub('LOGISTICS', '', buyernamecatching10)
	buyernamecatching12 = re.sub('SOLUTIONS', '', buyernamecatching11)

	while(z==0):
		listofbuyer = sheet2.cell(row = a, column = 1)
		buyerlist = listofbuyer.value
		if buyerlist == None:
			break
		buyerlistcatching = re.sub('PRIVATE' , '', buyerlist)
		buyerlistcatching2 = re.sub('LIMITED', '', buyerlistcatching)
		buyerlistcatching3 = re.sub('Private', '', buyerlistcatching2)
		buyerlistcatching4 = re.sub('Limited', '', buyerlistcatching3)
		buyerlistcatching5 = re.sub('CONSTRUCTION', '', buyerlistcatching4)
		buyerlistcatching6 = re.sub('ENTERPRISE', '', buyerlistcatching5)
		buyerlistcatching7 = re.sub('TRADERS', '', buyerlistcatching6)
		buyerlistcatching8 = re.sub('INDUSTRIES', '', buyerlistcatching7)
		buyerlistcatching9 = re.sub('SERVICES', '', buyerlistcatching8)
		buyerlistcatching10 = re.sub('TECHNOLOGIES', '', buyerlistcatching9)
		buyerlistcatching11 = re.sub('LOGISTICS', '', buyerlistcatching10)
		buyerlistcatching12 = re.sub('SOLUTIONS', '', buyerlistcatching11)		

		pmatch = int(SequenceMatcher(None, buyernamecatching12, buyerlistcatching12).ratio()*100)
		if(pmatch>percentagematch):
			percentagematch = pmatch
			buyer = str(buyerlist)
		a = int(a+1)
	
	if(percentagematch>=75):
		worksheet.write(Uu, buyer)		
		worksheet.write(Vv, (str(percentagematch) + '%'))
	else:
		worksheet.write(Uu, '')
		worksheet.write(Vv, '')
	b = int(b+1)
	x = int(x+1)
	print(x)

az = int(1)

print('A:')
while(az!=3762):
	Aa = str('A' + str(a))
	application = sheet.cell(row = az, column = 1)
	applicationvalue = application.value
	worksheet.write(Aa, applicationvalue)
	az = int(az+1)
	print(az)

z = int(1)

print('B:')
while(z!=3762):
	Bb = str('B' + str(z))
	applicant = sheet.cell(row = z, column = 2)
	applicantvalue = applicant.value
	worksheet.write(Bb, applicantvalue)
	z = int(z+1)
	print(z)

q = int(1)

print('C:')
while(q!=3762):
	Cc = str('C' + str(q))
	amount = sheet.cell(row = q, column = 3)
	amountvalue = amount.value
	worksheet.write(Cc, amountvalue)
	q = int(q+1)
	print(q)

print('D:')
r = int(1)
while(r!=3762):
	Dd = str('D' + str(r))
	narration = sheet.cell(row = r, column = 4)
	narrationvalue = narration.value
	worksheet.write(Dd, narrationvalue)
	r = int(r+1)
	print(r)

print('E:')
s = int(1)
while(s!=3762):
	Ee = str('E' + str(s))
	march = sheet.cell(row = s, column = 5)
	marchvalue = march.value
	worksheet.write(Ee, marchvalue)
	s = int(s+1)
	print(s)

print('F:')
t = int(1)
while(t!=3762):
	Ff = str('F' + str(t))
	april = sheet.cell(row = t, column = 6)
	aprilvalue = april.value
	worksheet.write(Ff, aprilvalue)
	t = int(t+1)
	print(t)

print('G:')
p = int(1)
while(p!=3762):
	Gg = str('G' + str(p))
	may = sheet.cell(row = p, column = 7)
	mayvalue = may.value
	worksheet.write(Gg, mayvalue)
	p = int(p+1)
	print(p)

print("H:")
l = int(1)
while(l!=3762):
	Hh = str('H' + str(l))
	june = sheet.cell(row = l, column = 8)
	junevalue = june.value
	worksheet.write(Hh, junevalue)
	l = int(l+1)
	print(l)

print("I:")
m = int(1)
while(m!=3762):
	Ii = str('I' + str(m))
	july = sheet.cell(row = m, column = 9)
	julyvalue = july.value
	worksheet.write(Ii, julyvalue)
	m = int(m+1)
	print(m)

print("J:")
m1 = int(1)
while(m1!=3762):
	Jj = str('J' + str(m1))
	august = sheet.cell(row = m1, column = 10)
	augustvalue = august.value
	worksheet.write(Jj, augustvalue)
	m1 = int(m1+1)
	print(m1)

print("K:")
m2 = int(1)
while(m2!=3762):
	Kk = str('K' + str(m2))
	september = sheet.cell(row = m2, column = 11)
	septembervalue = september.value
	worksheet.write(Kk, septembervalue)
	m2 = int(m2+1)
	print(m2)

print("L:")
m3 = int(1)
while(m3!=3762):
	Ll = str('L' + str(m3))
	october = sheet.cell(row = m3, column = 12)
	octobervalue = october.value
	worksheet.write(Ll, octobervalue)
	m3 = int(m3+1)
	print(m3)

print("M:")
m4 = int(1)
while(m4!=3762):
	Mm = str('M' + str(m4))
	november = sheet.cell(row = m4, column = 13)
	novembervalue = november.value
	worksheet.write(Mm, novembervalue)
	m4 = int(m4+1)
	print(m4)

print('MN:')
m5 = int(1)
while(m5!=3762):
	Nn = str('N' + str(m5))
	december = sheet.cell(row = m5, column = 14)
	decembervalue = december.value
	worksheet.write(Nn, decembervalue)
	m5 = int(m5+1)
	print(m5)

print("O:")
m6 = int(1)
while(m6!=3762):
	Oo = str('O' + str(m6))
	january = sheet.cell(row = m6, column = 15)
	januaryvalue = january.value
	worksheet.write(Oo, januaryvalue)
	m6 = int(m6+1)
	print(m6)

print("P:")
m7 = int(1)
while(m7!=3762):
	Pp = str('P' + str(m7))
	february = sheet.cell(row = m7, column = 16)
	februaryvalue = february.value
	worksheet.write(Pp, februaryvalue)
	m7 = int(m7+1)
	print(m7)

print("Q:")
m8 = int(1)
while(m8!=3762):
	Qq = str('Q' + str(m8))
	average = sheet.cell(row = m8, column = 17)
	averagevalue = average.value
	worksheet.write(Qq, averagevalue)
	m8 = int(m8+1)
	print(m8)

print("R:")
m9 = int(1)
while(m9!=3762):
	Rr = str('R' + str(m9))
	tag = sheet.cell(row = m9, column = 18)
	tagvalue = tag.value
	worksheet.write(Rr, tagvalue)
	m9 = int(m9+1)
	print(m9)

print("S:")
m10 = int(1)
while(m10!=3762):
	Ss = str('S' + str(m10))
	buyers1 = sheet.cell(row = m10, column = 19)
	buyers1value = buyers1.value
	worksheet.write(Ss, buyers1value)
	m10 = int(m10+1)
	print(m10)

print("T:")
m11 = int(1)
while(m11!=3762):
	Tt = str('T' + str(m11))
	dependancy = sheet.cell(row = m11, column = 20)
	dependancyvalue = dependancy.value
	worksheet.write(Tt, dependancyvalue)
	m11 = int(m11+1)
	print(m11)

workbook.close()