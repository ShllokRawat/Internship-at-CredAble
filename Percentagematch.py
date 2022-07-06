#ALL THE MODULES ARE IMPORTED HERE
from difflib import SequenceMatcher
import openpyxl
import xlsxwriter
import re
from datetime import datetime
#ALL THE MODULES ARE IMPORTED HERE

workbook = xlsxwriter.Workbook("Percentage Match.xlsx")
wb = openpyxl.load_workbook("C:\\Users\\shllo\\Desktop\\Internship\\Data\\Final Data\\BSA analysis report 28-06-2022 11.27.42.xlsx")
worksheet = workbook.add_worksheet()
wb.active = wb['Sheet1']
wb2 = openpyxl.load_workbook("C:\\Users\\shllo\\Desktop\\Internship\\Data\\Final Data\\Buyer list .xlsx")
wb2.active = wb2['Sheet1']
sheet2 = wb2.active
sheet = wb.active
x = int(2)
z = int(0)
b = int(1)
while(x!=3762):
	if(b == 1):
		worksheet.write('A1', 'Buyer')
		worksheet.write('B1', 'Percentage Match')
		worksheet.write('C1', 'Buyer Matched')
		b = int(b+1)
	Aa = str('A' + str(b))
	Bb = str('B' + str(b))
	Cc = str('C' + str(b))
	a = int(1)
	percentagematch = int(0)
	nameofbuyer = sheet.cell(row = x, column = 19)
	if(nameofbuyer.value == None):
		b = int(b+1)
		Aa = str('A' + str(b))
		Bb = str('B' + str(b))
		Cc = str('C' + str(b))
		x = int(x+1)
		buyername2 = sheet.cell(row = x, column = 19)
		if(buyername2.value == 'Buyer'):
			x = int(x+1)
			worksheet.write(Aa, 'Buyer')
			worksheet.write(Bb, 'Percentage Match')
			worksheet.write(Cc, 'Buyer Matched')
			b = int(b+1)
			continue
		elif(buyername2 == None):
			break
	buyername = str(nameofbuyer.value)
	buyernamecatching = re.sub('PRIVATE' , '', buyername)
	buyernamecatching2 = re.sub('LIMITED', '', buyernamecatching)
	buyernamecatching3 = re.sub('Private', '', buyernamecatching2)
	buyernamecatching4 = re.sub('Limited', '', buyernamecatching3)
	buyernamecatching5 = re.sub('CONSTRUCTION', '', buyernamecatching4)
	buyernamecatching6 = re.sub('ENTERPRISES', '', buyernamecatching5)
	buyernamecatching7 = re.sub('TRADERS', '', buyernamecatching6)
	buyernamecatching8 = re.sub('INDUSTRIES', '', buyernamecatching7)
	buyernamecatching9 = re.sub('SERVICES', '', buyernamecatching8)
	buyernamecatching10 = re.sub('TECHNOLOGIES', '', buyernamecatching9)
	buyernamecatching11 = re.sub('LOGISTICS', '', buyernamecatching10)
	buyernamecatching12 = re.sub('SOLUTIONS', '', buyernamecatching11)

	while(z==0):
		listofbuyer = sheet2.cell(row = a, column = 1)
		buyerlist = listofbuyer.value
		if buyerlist == None:
			break
		buyerlistcatching = re.sub('PRIVATE' , '', buyerlist)
		buyerlistcatching2 = re.sub('LIMITED', '', buyerlistcatching)
		buyerlistcatching3 = re.sub('Private', '', buyerlistcatching2)
		buyerlistcatching4 = re.sub('Limited', '', buyerlistcatching3)
		buyerlistcatching5 = re.sub('CONSTRUCTION', '', buyerlistcatching4)
		buyerlistcatching6 = re.sub('ENTERPRISE', '', buyerlistcatching5)
		buyerlistcatching7 = re.sub('TRADERS', '', buyerlistcatching6)
		buyerlistcatching8 = re.sub('INDUSTRIES', '', buyerlistcatching7)
		buyerlistcatching9 = re.sub('SERVICES', '', buyerlistcatching8)
		buyerlistcatching10 = re.sub('TECHNOLOGIES', '', buyerlistcatching9)
		buyerlistcatching11 = re.sub('LOGISTICS', '', buyerlistcatching10)
		buyerlistcatching12 = re.sub('SOLUTIONS', '', buyerlistcatching11)		

		pmatch = int(SequenceMatcher(None, buyernamecatching12, buyerlistcatching12).ratio()*100)
		if(pmatch>percentagematch):
			percentagematch = pmatch
			buyer = str(buyerlist)
		a = int(a+1)
	worksheet.write(Aa, buyername)
	if(percentagematch>=75):
		worksheet.write(Bb, buyer)		
		worksheet.write(Cc, (str(percentagematch) + '%'))
	else:
		worksheet.write(Bb, '')
		worksheet.write(Cc, '')
	b = int(b+1)
	x = int(x+1)
	print(x)

workbook.close()